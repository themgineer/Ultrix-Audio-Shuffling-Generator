"""
This creates a GUI that will easily generate names and port labels
for Audio Shuffling Sources for a Ross Video Ultrix.
"""

import os, sys
sys.path.insert(0, os.path.abspath("."))

from pathlib import Path

from PySide6 import QtCore, QtWidgets, QtGui
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.utils import exceptions

from UI.main_window_ui import Ui_mw_main

try:
    from ctypes import windll  # Only exists on Windows.

    MYAPPID = "themgineer.audio_shuffling_generator.0.0.8"
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(MYAPPID)
except ImportError:
    pass

class OutputEmpty(Exception):
    """Creates custom OutputEmpty Exception"""

class InputEmpty(Exception):
    """Creates custom InputEmpty Exception"""

class UnknownError(Exception):
    """Catches any other weird errors"""


class MainWindow(QtWidgets.QMainWindow, Ui_mw_main):

    wb = Workbook()
    ws = worksheet

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.group_view.setExclusive(True)
        self.settings = QtCore.QSettings("Anthony Thompson", "Ultrix Audio Shuffling Generator")

        if self.settings.value("viewMode"):
            self.set_color_scheme(self.settings.value("viewMode"))
        else:
            self.set_color_scheme("system")

        self.action_quit.triggered.connect(self.close)
        self.pb_quit.clicked.connect(self.close)

        self.le_sourceFile.textChanged.connect(self.autofill_output)
        self.le_sourceFile.textChanged.connect(lambda: self.load_sheet(self.le_sourceFile.text()))

        self.ck_start_index.checkStateChanged.connect(self.set_start_status)
        self.sb_start_index.valueChanged.connect(self.get_first_source_name)
        self.sb_end_index.valueChanged.connect(self.get_last_source_name)

        self.action_open.triggered.connect(self.open_source_dialog)
        self.action_light.triggered.connect(lambda: self.set_color_scheme("light"))
        self.action_dark.triggered.connect(lambda: self.set_color_scheme("dark"))
        self.action_system.triggered.connect(lambda: self.set_color_scheme("system"))

        self.pb_srcBrowse.clicked.connect(self.open_source_dialog)
        self.pb_outBrowse.clicked.connect(self.open_dest_dialog)

        self.action_guess.triggered.connect(self.guess_end_id)

        self.action_shuffle.triggered.connect(self.process_file)
        self.pb_shuffle.clicked.connect(self.process_file)

    @QtCore.Slot()
    def set_color_scheme(self, choice):
        if self.settings.value("viewMode") is not choice:
            self.settings.setValue("viewMode", choice)
        if choice == "light":
            QtGui.QGuiApplication.styleHints().setColorScheme(QtCore.Qt.ColorScheme.Light)
            self.action_light.setChecked(True)
        elif choice == "dark":
            QtGui.QGuiApplication.styleHints().setColorScheme(QtCore.Qt.ColorScheme.Dark)
            self.action_dark.setChecked(True)
        else:
            QtGui.QGuiApplication.styleHints().setColorScheme(QtCore.Qt.ColorScheme.Unknown)
            self.action_system.setChecked(True)

    @QtCore.Slot()
    def load_sheet(self, source_file):
        """ Slot attempts to load worksheet from source file """

        self.wb = Workbook()
        self.ws = worksheet

        if source_file:
            source_path = Path(source_file)
            try:
                self.wb = load_workbook(source_file)
                self.ws = self.wb["sources"]
                max_id = self.ws["A"][-1].value
                if type(max_id) is int:
                    self.sb_start_index.setMaximum(max_id)
                    self.sb_end_index.setMaximum(max_id)
                else:
                    raise TypeError
                self.show_status(f"Successfully loaded file: {source_path.name}")
            except exceptions.InvalidFileException:
                self.show_status(f"Cannot load file: {source_path.name}")
            except FileNotFoundError:
                self.show_status(f"File not found: {source_path.name}")
            except KeyError:
                self.show_status(f"{source_path.name} does not have a 'sources' worksheet.")
                self.le_sourceFile.clear()
            except TypeError:
                self.show_status(f"{source_path.name} does not appear to have any valid data")
                self.le_sourceFile.clear()

    @QtCore.Slot()
    def autofill_output(self):
        """ Generates output file name based on source file name """
        if self.le_sourceFile.text() == "":
            self.le_outputFile.setText("")
        else:
            source_path = Path(self.le_sourceFile.text())
            source_folder = source_path.parent
            file_name = source_path.stem
            file_ext = source_path.suffix
            output_path = Path.joinpath(source_folder, file_name + "_shuffled" + file_ext)
            self.le_outputFile.setText(str(output_path))

    @QtCore.Slot()
    def open_source_dialog(self):
        """ Opens a file dialog for existing files """
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Select a File",
            "",
            "Excel Files (*.xlsx)"
        )
        if filename:
            path = Path(filename)
            self.le_sourceFile.setText(str(path))

    @QtCore.Slot()
    def open_dest_dialog(self):
        """ Opens a file dialog for any files """
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Select a File",
            "",
            "Excel Files (*.xlsx)"
        )
        if filename:
            path = Path(filename)
            self.le_outputFile.setText(str(path))

    @QtCore.Slot()
    def set_start_status(self):
        """ Enables or disables Start Index spinbox based on state of checkbox """
        if self.ck_start_index.isChecked():
            self.sb_start_index.setEnabled(True)
        else:
            self.sb_start_index.setEnabled(False)

    @QtCore.Slot()
    def get_first_source_name(self):
        """ Attempts to find the source name of the line matching the set index """

        if self.sb_start_index.value() > self.sb_end_index.value():
            self.sb_end_index.setValue(self.sb_start_index.value())

        try:
            if self.le_sourceFile.text():
                first_source = self.ws['B'][self.sb_start_index.value() + 1].value # type: ignore
                self.show_status(f"First source: {first_source}")
        except TypeError:
            self.show_status("Invalid index")
        except IndexError:
            self.show_status("Index out of range")

    @QtCore.Slot()
    def get_last_source_name(self):
        """ Attempts to find the source name of the line matching the set index """

        if self.sb_start_index.value() > self.sb_end_index.value():
            self.sb_start_index.setValue(self.sb_end_index.value())

        try:
            if self.le_sourceFile.text():
                last_source = self.ws['B'][self.sb_end_index.value() + 1].value # type: ignore
                self.show_status(f"Last source: {last_source}")
        except TypeError:
            self.show_status("Invalid index")
        except IndexError:
            self.show_status("Index out of range")

    @QtCore.Slot()
    def process_file(self):
        """ Processes files """
        source_file = self.le_sourceFile.text()
        output_file = self.le_outputFile.text()
        start_id = self.sb_start_index.value()
        end_id = self.sb_end_index.value()
        channels = int(self.cb_channels.currentText())
        grouping = self.cb_grouping.currentText()
        leading_zero = self.ck_leadingZero.isChecked()

        # Initialize lists
        names = []
        ports = []

        group_dict = {"Mono": 1, "Stereo": 2, "Quad": 4, "Octo": 8}
        group_step = min(group_dict[grouping], channels)

        try:

            if source_file == "" or source_file is None:
                raise InputEmpty()

            if output_file == "" or output_file is None:
                raise OutputEmpty()

            if not Path(source_file).exists():
                raise FileNotFoundError()

            # Make sure sheet is loaded
            self.load_sheet(source_file)

            for row in self.ws.iter_rows(min_row=start_id + 2, # type: ignore
                                    max_row=end_id + 2,
                                    min_col=2,
                                    max_col=5,
                                    values_only = True):
                names.append(row[0])
                ports.append(row[3][:-8]) # type: ignore

            starting_index = self.ws['A'][-1].value + 1 # type: ignore

            shuffled_names = self.process_names(names, channels, group_step, leading_zero)
            shuffled_ports = self.process_ports(ports, channels, group_step)

            new_rows = self.create_new_rows(shuffled_names, shuffled_ports, starting_index)

            for row in new_rows:
                self.ws.append(row) # type: ignore

            # Checks to see if the Move Disconnect action is checked
            if self.action_move_disc.isChecked():
                self.move_disconnect()

            self.wb.save(output_file)

            message = "Output Successful!"

        except InputEmpty:
            message = "Input file not defined."
        except OutputEmpty:
            message = "Output file not defined."
        except FileNotFoundError:
            message = f"File not found: {Path(source_file).name}"
        except PermissionError:
            message = "Permission Denied: Excel file may still be open."
        except UnknownError:
            message = "Unknown error occurred."
        except AttributeError:
            message = "Unknown error with input file."

        self.show_status(message)

    QtCore.Slot()
    def guess_end_id(self):
        try:
            for row in self.ws.iter_rows(min_row=self.ws.min_row + 1, max_row=self.ws.max_row, max_col=self.ws.max_column):
                if type(row[5].value) is type(None) or "DANTE" in row[1].value.upper() or "MADI" in row[1].value.upper():
                    break

                last_source_row = row[0].value
            
            self.sb_end_index.setValue(last_source_row)
        except AttributeError:
            self.show_status("Cannot find rows in file. Is a valid file loaded?")

    def process_names(self, names, channels, group_step, leading_zero):
        """ Creates a list of names with channel numbers appended """

        output = []

        if leading_zero:
            if group_step == 1:
                for i in names:
                    for j in range(1, channels + 1):
                        output.append(f"{i} CH{j:02d}")
            else:
                for i in names:
                    for j in range(1, channels + 1, group_step):
                        output.append(f"{i} CH{j:02d}-{j+(group_step - 1):02d}")
        else:
            if group_step == 1:
                for i in names:
                    for j in range(1, channels + 1):
                        output.append(f"{i} CH{j}")
            else:
                for i in names:
                    for j in range(1, channels + 1, group_step):
                        output.append(f"{i} CH{j}-{j+(group_step - 1)}")

        return output

    def process_ports(self, ports, channels, group_step):
        """ Creates a list of lists with channel ports """

        output = []

        for port in ports:
            for ch_range in range(1, channels + 1, group_step):
                temp = []
                for _ in range(1, channels + 1, group_step):
                    for ch in range(ch_range, ch_range + group_step):
                        temp.append(f"{port}.audio.ch{ch}")
                output.append(temp)
        return output

    def create_new_rows(self, names, ports, index):
        """ Creates rows to add to Excel sheet """
        output = []

        for i, name in enumerate(names):
            temp = [index, name, "", "", ""]
            temp.extend(ports[i])
            output.append(temp)
            index += 1

        return output

    def show_status(self, message):
        """ Updates status bar message """
        self.statusbar.showMessage(message, timeout=5000)
    
    def move_disconnect(self):
        """ Finds the first instance of a row containing "DISC" in the name and moves it to the bottom of the sheet
            It then renumbers the sheet to make sure the ID numbers are sequential """
        disc_row = None
        
        for row in self.ws.iter_rows(min_row=self.ws.min_row + 1, max_row=self.ws.max_row, max_col=self.ws.max_column):
            if "DISC" in row[1].value.upper():
                disc_row = row
                break

        if disc_row:
            self.ws.delete_rows(disc_row[0].row)
            self.ws.append(disc_row)

            for row in self.ws.iter_rows(min_row=self.ws.min_row + 1, max_row=self.ws.max_row, max_col=1):
                row[0].value = row[0].row - 2


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()

    sys.exit(app.exec())
